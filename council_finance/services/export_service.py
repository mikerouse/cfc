"""
Export Service
Provides data export functionality in multiple formats (CSV, XLSX, PDF, PNG)
"""

import io
import csv
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.staticfiles import finders

# Third-party imports
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logging.warning("openpyxl not installed. XLSX export will not be available.")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("reportlab not installed. PDF export will not be available.")

try:
    from PIL import Image, ImageDraw, ImageFont
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    PNG_AVAILABLE = True
except ImportError:
    PNG_AVAILABLE = False
    logging.warning("PIL/matplotlib not installed. PNG export will not be available.")

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting data in various formats"""
    
    def __init__(self):
        self.supported_formats = ['csv']
        if XLSX_AVAILABLE:
            self.supported_formats.append('xlsx')
        if PDF_AVAILABLE:
            self.supported_formats.append('pdf')
        if PNG_AVAILABLE:
            self.supported_formats.append('png')
    
    def export_leaderboard(
        self,
        data: Dict[str, Any],
        format: str,
        filename: Optional[str] = None
    ) -> HttpResponse:
        """Export leaderboard data in the specified format"""
        
        if format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format}. Supported: {self.supported_formats}")
        
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"leaderboard_{data.get('category', 'data')}_{timestamp}"
        
        if format == 'csv':
            return self._export_csv(data, filename)
        elif format == 'xlsx':
            return self._export_xlsx(data, filename)
        elif format == 'pdf':
            return self._export_pdf(data, filename)
        elif format == 'png':
            return self._export_png(data, filename)
    
    def _export_csv(self, data: Dict[str, Any], filename: str) -> HttpResponse:
        """Export data as CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            f"{data.get('category_name', 'Leaderboard')} - {data.get('year', 'All Years')}",
        ])
        if data.get('per_capita'):
            writer.writerow(['Per Capita Values'])
        writer.writerow([])  # Empty row
        
        # Write column headers
        entries = data.get('entries', [])
        if entries and isinstance(entries[0], dict):
            if 'username' in entries[0]:  # Contributors
                headers = ['Rank', 'Username', 'Points', 'Badge']
            else:  # Financial
                headers = ['Rank', 'Council', 'Type', 'Nation', 'Value (£)']
                if data.get('per_capita'):
                    headers.extend(['Population', 'Per Capita (£)'])
            writer.writerow(headers)
            
            # Write data rows
            for entry in entries:
                if 'username' in entry:
                    row = [
                        entry.get('rank'),
                        entry.get('username'),
                        entry.get('points'),
                        entry.get('badge', '')
                    ]
                else:
                    row = [
                        entry.get('rank'),
                        entry.get('council_name'),
                        entry.get('council_type', ''),
                        entry.get('council_nation', ''),
                        f"{entry.get('value', 0):,.2f}"
                    ]
                    if data.get('per_capita'):
                        row.extend([
                            f"{entry.get('population', 0):,}",
                            f"{entry.get('per_capita_value', 0):,.2f}"
                        ])
                writer.writerow(row)
        
        return response
    
    def _export_xlsx(self, data: Dict[str, Any], filename: str) -> HttpResponse:
        """Export data as XLSX"""
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX export")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = data.get('category_name', 'Leaderboard')[:31]  # Excel limit
        
        # Set column widths
        ws.column_dimensions['A'].width = 8   # Rank
        ws.column_dimensions['B'].width = 35  # Name
        ws.column_dimensions['C'].width = 20  # Type
        ws.column_dimensions['D'].width = 15  # Nation
        ws.column_dimensions['E'].width = 15  # Value
        
        # Title
        title = f"{data.get('category_name', 'Leaderboard')} - {data.get('year', 'All Years')}"
        ws.merge_cells('A1:E1')
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Subtitle if per capita
        row = 2
        if data.get('per_capita'):
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = 'Per Capita Values'
            ws[f'A{row}'].font = Font(size=12, italic=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        row += 1  # Empty row
        
        # Headers
        entries = data.get('entries', [])
        if entries and isinstance(entries[0], dict):
            if 'username' in entries[0]:  # Contributors
                headers = ['Rank', 'Username', 'Points', 'Badge', '']
            else:  # Financial
                headers = ['Rank', 'Council', 'Type', 'Nation', 'Value (£)']
                if data.get('per_capita'):
                    headers.extend(['Population', 'Per Capita (£)'])
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 15
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            row += 1
            
            # Data rows
            for entry in entries:
                if 'username' in entry:
                    values = [
                        entry.get('rank'),
                        entry.get('username'),
                        entry.get('points'),
                        entry.get('badge', ''),
                        ''
                    ]
                else:
                    values = [
                        entry.get('rank'),
                        entry.get('council_name'),
                        entry.get('council_type', ''),
                        entry.get('council_nation', ''),
                        entry.get('value', 0)
                    ]
                    if data.get('per_capita'):
                        values.extend([
                            entry.get('population', 0),
                            entry.get('per_capita_value', 0)
                        ])
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    # Format numbers
                    if col >= 5 and isinstance(value, (int, float)):
                        if col == 5 or (data.get('per_capita') and col == 7):
                            cell.number_format = '£#,##0.00'
                        elif data.get('per_capita') and col == 6:
                            cell.number_format = '#,##0'
                    
                    # Highlight top 3
                    if col == 1 and value <= 3:
                        fill_color = {
                            1: "FFD700",  # Gold
                            2: "C0C0C0",  # Silver
                            3: "CD7F32"   # Bronze
                        }.get(value, "FFFFFF")
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                row += 1
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        
        return response
    
    def _export_pdf(self, data: Dict[str, Any], filename: str) -> HttpResponse:
        """Export data as PDF"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        # Create PDF
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1F4E79'),
            spaceAfter=30,
            alignment=1  # Center
        )
        title = f"{data.get('category_name', 'Leaderboard')} - {data.get('year', 'All Years')}"
        elements.append(Paragraph(title, title_style))
        
        if data.get('per_capita'):
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=14,
                textColor=colors.gray,
                spaceAfter=20,
                alignment=1
            )
            elements.append(Paragraph("Per Capita Values", subtitle_style))
        
        elements.append(Spacer(1, 20))
        
        # Table data
        entries = data.get('entries', [])
        if entries and isinstance(entries[0], dict):
            if 'username' in entries[0]:  # Contributors
                table_data = [['Rank', 'Username', 'Points', 'Badge']]
                col_widths = [50, 200, 80, 100]
            else:  # Financial
                headers = ['Rank', 'Council', 'Type', 'Value (£)']
                col_widths = [50, 250, 100, 100]
                if data.get('per_capita'):
                    headers.extend(['Population', 'Per Capita (£)'])
                    col_widths.extend([80, 100])
                table_data = [headers]
            
            # Add data rows
            for entry in entries[:50]:  # Limit to 50 for PDF
                if 'username' in entry:
                    row = [
                        str(entry.get('rank')),
                        entry.get('username', ''),
                        str(entry.get('points', 0)),
                        entry.get('badge', '')
                    ]
                else:
                    row = [
                        str(entry.get('rank')),
                        entry.get('council_name', ''),
                        entry.get('council_type', ''),
                        f"£{entry.get('value', 0):,.0f}"
                    ]
                    if data.get('per_capita'):
                        row.extend([
                            f"{entry.get('population', 0):,}",
                            f"£{entry.get('per_capita_value', 0):,.2f}"
                        ])
                table_data.append(row)
            
            # Create table
            table = Table(table_data, colWidths=col_widths)
            
            # Style table
            table_style = TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Rank column
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),  # Numeric columns
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ])
            
            # Highlight top 3
            for i in range(1, min(4, len(table_data))):
                color = {
                    1: colors.HexColor('#FFD700'),  # Gold
                    2: colors.HexColor('#C0C0C0'),  # Silver
                    3: colors.HexColor('#CD7F32')   # Bronze
                }.get(i)
                if color:
                    table_style.add('BACKGROUND', (0, i), (0, i), color)
            
            table.setStyle(table_style)
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return response
    
    def _export_png(self, data: Dict[str, Any], filename: str) -> HttpResponse:
        """Export data as PNG chart"""
        if not PNG_AVAILABLE:
            raise ImportError("PIL and matplotlib are required for PNG export")
        
        # Create figure
        fig, ax = plt.subplots(figsize=(12, 8))
        
        entries = data.get('entries', [])[:10]  # Top 10 for visualization
        
        if entries and isinstance(entries[0], dict):
            if 'username' in entries[0]:  # Contributors
                names = [e.get('username', '') for e in entries]
                values = [e.get('points', 0) for e in entries]
                ylabel = 'Points'
            else:  # Financial
                names = [e.get('council_name', '') for e in entries]
                if data.get('per_capita'):
                    values = [e.get('per_capita_value', 0) for e in entries]
                    ylabel = 'Value per Capita (£)'
                else:
                    values = [e.get('value', 0) for e in entries]
                    ylabel = 'Value (£)'
            
            # Create bar chart
            bars = ax.bar(range(len(names)), values)
            
            # Color top 3
            if len(bars) > 0:
                bars[0].set_color('#FFD700')  # Gold
            if len(bars) > 1:
                bars[1].set_color('#C0C0C0')  # Silver
            if len(bars) > 2:
                bars[2].set_color('#CD7F32')  # Bronze
            
            # Customize chart
            ax.set_xticks(range(len(names)))
            ax.set_xticklabels(names, rotation=45, ha='right')
            ax.set_ylabel(ylabel)
            ax.set_title(f"{data.get('category_name', 'Leaderboard')} - {data.get('year', 'All Years')}")
            
            # Add value labels on bars
            for i, (bar, value) in enumerate(zip(bars, values)):
                height = bar.get_height()
                if 'username' in entries[0]:
                    label = f'{int(value):,}'
                else:
                    label = f'£{value:,.0f}'
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       label, ha='center', va='bottom')
            
            plt.tight_layout()
            
            # Save to response
            response = HttpResponse(content_type='image/png')
            response['Content-Disposition'] = f'attachment; filename="{filename}.png"'
            plt.savefig(response, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            
            return response
        
        # Empty response if no data
        response = HttpResponse(content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="{filename}.png"'
        return response